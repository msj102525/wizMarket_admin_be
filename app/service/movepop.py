import pymysql
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime
import os, time
from tqdm import tqdm
from app.db.connect import get_db_connection, close_connection


def crawl_keyword(keyword: str, connection, insert_count):
    driver_path = os.path.join(os.path.dirname(__file__), "../", "drivers", "chromedriver.exe")
    options = webdriver.ChromeOptions()
    options.add_argument('--headless')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')

    driver = webdriver.Chrome(service=Service(driver_path), options=options)

    try:
        driver.get('https://sg.sbiz.or.kr/godo/index.sg')

        radio_button = driver.find_element(By.ID, 'icon_list2_2')
        driver.execute_script("arguments[0].click();", radio_button)

        search_box = driver.find_element(By.ID, 'searchAddress')
        search_box.send_keys(keyword)
        search_box.send_keys(Keys.RETURN)

        last_keyword = keyword.split()[-1]

        print(f"마지막 동 단위 단어 : {last_keyword}")
        time.sleep(10)

        try:
            WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.XPATH, "//div[@class='cell']"))
            )
            div_elements = driver.find_elements(By.XPATH, "//div[@class='cell']")
            div_content = None
            for div_element in div_elements:
                try:
                    span_element = div_element.find_element(By.TAG_NAME, 'span')
                    if span_element.text == last_keyword:
                        div_content = div_element.get_attribute('innerHTML')
                        break
                except Exception as e:
                    print(f"Error accessing span element: {e}")
        except Exception as e:
            print(f"Timeout or error waiting for div element: {e}")
            div_content = None

        if div_content:
            data = parse_html(div_content)
        else:
            data = {
                'location': None,
                'business': None,
                'person': None,
                'price': None,
                'wrcppl': None,
                'earn': None,
                'cnsmp': None,
                'hhCnt': None,
                'rsdppl': None
            }

        year_month = datetime.now().strftime("%Y%m")
        insert_record('movepopdata', connection, insert_count, keyword=keyword, yearmonth=year_month, **data)

    except Exception as e:
        print(f'Error: {e}')
        raise
    finally:
        driver.quit()


def parse_html(html_content):
    from bs4 import BeautifulSoup
    soup = BeautifulSoup(html_content, 'html.parser')
    data = {
        'location': soup.find('span').text,
        'business': soup.find('strong', {'data-name': 'business'}).text,
        'person': soup.find('strong', {'data-name': 'person'}).text,
        'price': soup.find('strong', {'data-name': 'price'}).text,
        'wrcppl': soup.find('strong', {'data-name': 'wrcppl'}).text,
        'earn': soup.find('strong', {'data-name': 'earn'}).text,
        'cnsmp': soup.find('strong', {'data-name': 'cnsmp'}).text,
        'hhCnt': soup.find('strong', {'data-name': 'hhCnt'}).text,
        'rsdppl': soup.find('strong', {'data-name': 'rsdppl'}).text
    }
    return data


def insert_record(table_name: str, connection, insert_count, **kwargs):
    try:
        with connection.cursor() as cursor:
            columns = ', '.join(kwargs.keys())
            placeholders = ', '.join(['%s'] * len(kwargs))
            values = tuple(kwargs.values())
            sql = f"INSERT INTO {table_name} ({columns}) VALUES ({placeholders})"
            cursor.execute(sql, values)
        connection.commit()
        print(f"{insert_count}번째 insert 성공")
    except pymysql.IntegrityError as e:
        if e.args[0] == 1062:  # Duplicate entry error code
            print(f"Duplicate entry found for {kwargs}. Skipping...")
        else:
            print(f"Error inserting data: {e}")
            connection.rollback()
            raise e
    except Exception as e:
        print(f"Error inserting data: {e}")
        connection.rollback()
        raise e


def read_keywords_from_excel(file_path):
    workbook = load_workbook(filename=file_path)
    sheet = workbook.active
    keywords = [row[0].value for row in sheet.iter_rows(min_row=1) if row[0].value]
    return keywords


def process_file(file_path):
    connection = get_db_connection()
    insert_count = 0
    try:
        keywords = read_keywords_from_excel(file_path)
        for keyword in tqdm(keywords, desc=f"Processing {os.path.basename(file_path)}"):  # tqdm을 사용하여 진행 상황 표시
            # DB에서 해당 키워드가 이미 있는지 확인
            with connection.cursor() as cursor:
                cursor.execute("SELECT 1 FROM movepopdata WHERE keyword = %s", (keyword,))
                result = cursor.fetchone()
                if result:
                    print(f"Keyword {keyword} already exists. Skipping...")
                    continue

            print(f"Processing keyword: {keyword}")
            insert_count += 1
            crawl_keyword(keyword, connection, insert_count)
    finally:
        close_connection(connection)
    print(f"Finished processing {os.path.basename(file_path)}")


def process_keywords_from_excel():
    print("Starting to process keywords from Excel")
    directory = "C:/formovedata"
    excel_files = [os.path.join(directory, f"SplitFile_{i}.xlsx") for i in [6, 7, 8]]
    # excel_files = [os.path.join(directory, f"list.xlsx")] 밑의 쓰레드 숫자 1로 바꾸기

    with ThreadPoolExecutor(max_workers=3) as executor:
        executor.map(process_file, excel_files)

    print("Finished processing all files")

