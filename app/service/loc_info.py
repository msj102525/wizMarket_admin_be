from fastapi import HTTPException
import pymysql
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from concurrent.futures import ThreadPoolExecutor
from datetime import date, datetime
import os, time
from tqdm import tqdm
import sys
from app.crud.loc_info import *
import pandas as pd

from app.crud.loc_store import (
    select_local_store_sub_distirct_id_by_store_business_number as crud_select_local_store_sub_distirct_id_by_store_business_number,
)
from app.crud.statistics import (
    select_nationwide_jscore_by_stat_item_id_and_sub_district_id as crud_select_nationwide_jscore_by_stat_item_id_and_sub_district_id,
    select_state_item_id as crud_select_state_item_id,
)
from app.schemas.loc_store import LocalStoreSubdistrict
from app.schemas.statistics import DataRefDateSubDistrictName, LocInfoStatisticsDataRefOutput, LocInfoStatisticsOutput


async def filter_location_info(filters: dict):
    # 필터링 로직: 필요하면 여기서 추가적인 필터 처리를 할 수 있습니다.
    filtered_locations = get_filtered_locations(filters)
    all_corr = get_all_corr()

    # 필요한 항목들로 DataFrame 생성 (SALES와 다른 변수들)
    df_all = pd.DataFrame(
        all_corr,
        columns=[
            "SALES",
            "SHOP",
            "MOVE_POP",
            "WORK_POP",
            "INCOME",
            "SPEND",
            "HOUSE",
            "RESIDENT",
        ],
    )

    # 전체 상관분석 수행
    all_corr_matrix = df_all.corr()

    # 지역 내 상관 분석
    filter_corr = get_filter_corr(filters)

    # 지역 내 분석 수행
    df_filter = pd.DataFrame(filter_corr)

    filter_corr_matrix = df_filter.groupby("DISTRICT_NAME")[
        [
            "SALES",
            "SHOP",
            "MOVE_POP",
            "WORK_POP",
            "INCOME",
            "SPEND",
            "HOUSE",
            "RESIDENT",
        ]
    ].corr()

    filter_corr_matrix = filter_corr_matrix.reset_index().to_dict(orient="records")

    # 필요 시 추가적인 비즈니스 로직을 처리할 수 있음
    return filtered_locations, all_corr_matrix, filter_corr_matrix




###### 크롤링 #########
sys.path.append(os.path.join(os.path.dirname(__file__), "../../"))
from app.db.connect import get_db_connection, close_connection

# 1. 가지고 있는 지역 id값 모두 조회
def fetch_all_region_id():
    all_region_list = get_all_region_id()
    return all_region_list

# Global WebDriver instance
global_driver = None


def setup_global_driver():
    global global_driver
    if global_driver is None:
        options = Options()
        options.add_argument("--start-fullscreen")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")

        # WebDriver Manager를 이용해 ChromeDriver 자동 관리
        service = Service(ChromeDriverManager().install())
        global_driver = webdriver.Chrome(service=service, options=options)

    return global_driver


def crawl_keyword(region_data, connection, insert_count):
     # 글로벌 드라이버 사용
    global global_driver
    driver = setup_global_driver()

    try:
        driver.get("https://sg.sbiz.or.kr/godo/index.sg")

        # 페이지 로드 완료 확인 코드 추가
        WebDriverWait(driver, 30).until(
            lambda d: d.execute_script("return document.readyState") == "complete"
        )

        radio_button = driver.find_element(By.ID, "icon_list2_2")
        driver.execute_script("arguments[0].click();", radio_button)

        search_box = driver.find_element(By.ID, "searchAddress")
        search_box.send_keys(region_data['keyword'])
        search_box.send_keys(Keys.RETURN)

        last_keyword = region_data['keyword'].split()[-1]

        time.sleep(0.3)

        try:
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.XPATH, "//div[@class='cell']"))
            )
            div_elements = driver.find_elements(By.XPATH, "//div[@class='cell']")
            div_content = None
            for div_element in div_elements:
                try:
                    span_element = div_element.find_element(By.TAG_NAME, "span")
                    if span_element.text == last_keyword:
                        div_content = div_element.get_attribute("innerHTML")
                        break
                except Exception as e:
                    print(f"Error accessing span element: {e}")
        except Exception as e:
            print(f"Timeout or error waiting for div element: {e}")
            div_content = None

        if div_content:
            data = parse_html(div_content)
        else:
            data = {
                "shop": None,
                "move_pop": None,
                "sales": None,
                "work_pop": None,
                "income": None,
                "spend": None,
                "house": None,
                "resident": None,
            }
        print(data)
        year_month = datetime(2024, 10, 2).date()
        try:
            with connection.cursor() as cursor:
                insert_record(
                    "loc_info",
                    connection,
                    insert_count,
                    city_id=region_data['city_id'],
                    district_id=region_data['district_id'],
                    sub_district_id=region_data['sub_district_id'],
                    y_m=year_month,
                    **data,
                )
                connection.commit()
                    
        except Exception as e:
            print(f"Error inserting or updating data: {e}")
            connection.rollback()
            raise e

    except Exception as e:
        print(f"Error: {e}")
        raise
    finally:
        driver.quit()


def parse_html(html_content):
    from bs4 import BeautifulSoup

    def clean_value(value):
        if value:
            # 쉼표와 단위 제거
            value = value.replace(",", "").replace("만원", "").replace("명", "").strip()
        return value

    soup = BeautifulSoup(html_content, "html.parser")
    data = {
        "shop": clean_value(soup.find("strong", {"data-name": "business"}).text if soup.find("strong", {"data-name": "business"}) else None),
        "move_pop": clean_value(soup.find("strong", {"data-name": "person"}).text if soup.find("strong", {"data-name": "person"}) else None),
        "sales": clean_value(soup.find("strong", {"data-name": "price"}).text if soup.find("strong", {"data-name": "price"}) else None),
        "work_pop": clean_value(soup.find("strong", {"data-name": "wrcppl"}).text if soup.find("strong", {"data-name": "wrcppl"}) else None),
        "income": clean_value(soup.find("strong", {"data-name": "earn"}).text if soup.find("strong", {"data-name": "earn"}) else None),
        "spend": clean_value(soup.find("strong", {"data-name": "cnsmp"}).text if soup.find("strong", {"data-name": "cnsmp"}) else None),
        "house": clean_value(soup.find("strong", {"data-name": "hhCnt"}).text if soup.find("strong", {"data-name": "hhCnt"}) else None),
        "resident": clean_value(soup.find("strong", {"data-name": "rsdppl"}).text if soup.find("strong", {"data-name": "rsdppl"}) else None),
    }
    return data



def insert_record(table_name: str, connection, insert_count, city_id, district_id, sub_district_id, **kwargs):
    try:
        with connection.cursor() as cursor:
            # city_id, district_id, sub_district_id는 고정 컬럼
            columns = "city_id, district_id, sub_district_id, " + ", ".join(kwargs.keys())
            placeholders = "%s, %s, %s, " + ", ".join(["%s"] * len(kwargs))
            values = (city_id, district_id, sub_district_id) + tuple(kwargs.values())
            
            sql = f"INSERT INTO {table_name} ({columns}) VALUES ({placeholders})"
            cursor.execute(sql, values)
        connection.commit()
        print(f"{insert_count}번째 insert 성공")
    except pymysql.IntegrityError as e:
        if e.args[0] == 1062:  # Duplicate entry error code
            print(f"Duplicate entry found for {kwargs}. Skipping...")
        else:
            print(f"Error inserting data: {e}")
            connection.rollback()
            raise e
    except Exception as e:
        print(f"Error inserting data: {e}")
        connection.rollback()
        raise e


# 5. 엑셀 파일에서 값 꺼내서 키워드 화, 매핑
def read_keywords_from_excel(file_path, all_region_list):
    workbook = load_workbook(filename=file_path)
    sheet = workbook.active
    keywords = []

    # 엑셀 파일에서 시도명, 시군구명, 읍면동명을 읽어와 매칭
    for row in sheet.iter_rows(min_row=1):  # 1행부터 시작
        city_name = row[0].value  # 시도명
        district_name = row[1].value  # 시군구명
        sub_district_name = row[2].value  # 읍면동명

        if city_name and district_name and sub_district_name:
            # all_region_list에서 매칭되는 id 값을 찾음
            matching_region = next(
                (region for region in all_region_list 
                 if region['city_name'] == city_name 
                 and region['district_name'] == district_name 
                 and region['sub_district_name'] == sub_district_name),
                None
            )
            if matching_region:
                # 매칭된 지역의 id 값을 추출하고 리스트에 저장
                keywords.append({
                    'city_id': matching_region['city_id'],
                    'district_id': matching_region['district_id'],
                    'sub_district_id': matching_region['sub_district_id'],
                    'keyword': f"{city_name} {district_name} {sub_district_name}"
                })
            else:
                print(f"Region not found for: {city_name} {district_name} {sub_district_name}")
    return keywords



def process_file(file_path, all_region_list):
    connection = get_db_connection()
    insert_count = 0

    try:
        # all_region_list를 전달하여 지역 ID 값과 매핑된 키워드를 얻음
        keywords = read_keywords_from_excel(file_path, all_region_list)

        for keyword_data in tqdm(
            keywords, desc=f"Processing {os.path.basename(file_path)}"
        ):  # tqdm을 사용하여 진행 상황 표시
            # 각 키워드에 대해 무조건 새로운 데이터를 삽입하거나 갱신
            insert_count += 1
            
            # keyword_data는 city_id, district_id, sub_district_id, keyword를 포함
            crawl_keyword(keyword_data, connection, insert_count)

    finally:
        close_connection(connection)
    print(f"Finished processing {os.path.basename(file_path)}")



def process_keywords_from_excel():
    print("Starting to process keywords from Excel")
    all_region_list = fetch_all_region_id()
    directory = "C:/formovedata"
    excel_files = [os.path.join(directory, f"SplitFile_{i}.xlsx") for i in [1, 2, 3, 4, 5]]
    # excel_files = [os.path.join(directory, f"list.xlsx")]

    with ThreadPoolExecutor(max_workers=5) as executor:
        executor.map(lambda file: process_file(file, all_region_list), excel_files)

    print("Finished processing all files")


if __name__=="__main__":
    process_keywords_from_excel()
    # fetch_all_region_id()
    # read_keywords_from_excel("C:\\formovedata\\SplitFile_1.xlsx")


def select_report_loc_info_by_store_business_number(
    store_business_id: str,
) -> LocInfoStatisticsDataRefOutput:

    try:
        local_store_sub_district_data: LocalStoreSubdistrict = (
            crud_select_local_store_sub_distirct_id_by_store_business_number(
                store_business_id
            )
        )

        sub_district_id = local_store_sub_district_data.get("SUB_DISTRICT_ID")
        sub_district_name = local_store_sub_district_data.get("SUB_DISTRICT_NAME")

        # sub_district_id가 None일 경우 오류 처리
        if sub_district_id is None:
            raise HTTPException(status_code=500, detail="Sub-district ID not found.")

        # 각 통계 항목 ID 조회 및 출력
        mz_pupulation_stat_item_id: int = crud_select_state_item_id(
            "population", "mz_population"
        )
        # print(f"mz_pupulation_stat_item_id: {mz_pupulation_stat_item_id}")

        resident_stat_item_id: int = crud_select_state_item_id("loc_info", "resident")
        # print(f"resident_stat_item_id: {resident_stat_item_id}")

        move_pop_stat_item_id: int = crud_select_state_item_id("loc_info", "move_pop")
        # print(f"move_pop_stat_item_id: {move_pop_stat_item_id}")

        house_stat_item_id: int = crud_select_state_item_id("loc_info", "house")
        # print(f"house_stat_item_id: {house_stat_item_id}")

        shop_stat_item_id: int = crud_select_state_item_id("loc_info", "shop")
        # print(f"shop_stat_item_id: {shop_stat_item_id}")

        income_stat_item_id: int = crud_select_state_item_id("loc_info", "income")
        # print(f"income_stat_item_id: {income_stat_item_id}")

        spend_stat_item_id: int = crud_select_state_item_id("loc_info", "spend")
        # print(f"spend_stat_item_id: {spend_stat_item_id}")

        sales_stat_item_id: int = crud_select_state_item_id("loc_info", "sales")
        # print(f"sales_stat_item_id: {sales_stat_item_id}")

        # print(f"loc_info: {store_business_id}")
        # print(f"loc_info: {sub_district_id}")

        # 각 통계 데이터 조회 및 처리
        mz_population_jscore: float = round(
            crud_select_nationwide_jscore_by_stat_item_id_and_sub_district_id(
                mz_pupulation_stat_item_id, sub_district_id
            ).get("J_SCORE"),
            1,
        )
        shop_jscore: float = round(
            crud_select_nationwide_jscore_by_stat_item_id_and_sub_district_id(
                shop_stat_item_id, sub_district_id
            ).get("J_SCORE"),
            1,
        )
        move_pop_jscore: float = round(
            crud_select_nationwide_jscore_by_stat_item_id_and_sub_district_id(
                move_pop_stat_item_id, sub_district_id
            ).get("J_SCORE"),
            1,
        )
        resident_jscore: float = round(
            crud_select_nationwide_jscore_by_stat_item_id_and_sub_district_id(
                resident_stat_item_id, sub_district_id
            ).get("J_SCORE"),
            1,
        )
        house_jscore: float = round(
            crud_select_nationwide_jscore_by_stat_item_id_and_sub_district_id(
                house_stat_item_id, sub_district_id
            ).get("J_SCORE"),
            1,
        )
        income_jscore: float = round(
            crud_select_nationwide_jscore_by_stat_item_id_and_sub_district_id(
                income_stat_item_id, sub_district_id
            ).get("J_SCORE"),
            1,
        )
        spend_jscore: float = round(
            crud_select_nationwide_jscore_by_stat_item_id_and_sub_district_id(
                spend_stat_item_id, sub_district_id
            ).get("J_SCORE"),
            1,
        )
        sales_jscore: float = round(
            crud_select_nationwide_jscore_by_stat_item_id_and_sub_district_id(
                sales_stat_item_id, sub_district_id
            ).get("J_SCORE"),
            1,
        )
        data_ref_date: date = (
            crud_select_nationwide_jscore_by_stat_item_id_and_sub_district_id(
                mz_pupulation_stat_item_id, sub_district_id
            ).get("REF_DATE")
        )

        data_ref = DataRefDateSubDistrictName(
            sub_district_name = sub_district_name,
            reference_date = data_ref_date
        )

        # print(data_ref)

        # 각 stat_item_id와 jscore 출력
        # print(f"mz_population_jscore: {mz_population_jscore}")
        # print(f"shop_jscore: {shop_jscore}")
        # print(f"move_pop_jscore: {move_pop_jscore}")
        # print(f"resident_jscore: {resident_jscore}")
        # print(f"house_jscore: {house_jscore}")
        # print(f"income_jscore: {income_jscore}")
        # print(f"spend_jscore: {spend_jscore}")
        # print(f"sales_jscore: {sales_jscore}")

        loc_info_report_data = LocInfoStatisticsOutput(
            mz_population_jscore=mz_population_jscore,
            shop_jscore=shop_jscore,
            move_pop_jscore=move_pop_jscore,
            resident_jscore=resident_jscore,
            house_jscore=house_jscore,
            income_jscore=income_jscore,
            spend_jscore=spend_jscore,
            sales_jscore=sales_jscore,
        )

        loc_info_report_output = LocInfoStatisticsDataRefOutput(
            loc_info = loc_info_report_data,
            data_ref = data_ref
        )

        return loc_info_report_output

    except HTTPException as http_ex:
        raise http_ex
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"An error occurred: {str(e)}")


